"""
railway_data_preparation.py
===========================
STRATEGIC + OPERATIONAL data loaders for the railway platform.

Two independent business domains are loaded here but **never merged**:

  * Strategic  : raw_data/railways.xlsx
                 -> railway_demand_history.csv  (the Step-2 deliverable)
  * Operational: raw_data/railway_stock_summary.xlsx
                 -> in-memory frame consumed later by railway_operational_analysis.py

Design notes
------------
* The operational workbook has a CORRUPT embedded stylesheet, so openpyxl / pandas
  cannot open it. We parse it directly from the OOXML (zip + XML) -- the same
  technique validated during Phase-1 discovery.
* The strategic workbook opens cleanly with openpyxl; we read it by the *positional*
  column coordinates recorded in railway_config (discovered in Phase 1), so the
  loader scales unchanged from 59 -> 907 -> N rows.
* No Walmart code is imported or touched.
"""

from __future__ import annotations

import re
import zipfile
from xml.etree import ElementTree as ET

import numpy as np
import pandas as pd
import openpyxl

from railway import railway_config as cfg
from railway.ingestion import excel_reader, pl_normalization   # consolidated I/O (Phase A)

_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

# Final column order for railway_demand_history.csv (locked spec)
DEMAND_HISTORY_FIELDS = [
    "PL_Code", "Description",
    "FY2020_21", "FY2022_23", "FY2023_24", "FY2024_25", "FY2025_26",
    "AAC", "EAR_Qty", "Pending_Supply", "Current_Stock", "Unit_Cost",
]


# ----------------------------------------------------------------------
# small value-cleaning helpers
# ----------------------------------------------------------------------
def _to_num(v):
    """Coerce a cell to float, or NaN if not numeric ('--NA--', '', None, text)."""
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s == "" or s.upper() in {"--NA--", "NA", "N/A", "NIL", "-"}:
        return np.nan
    try:
        return float(s)
    except ValueError:
        return np.nan


def _norm_pl(v):
    """Normalise a P.L. code to a clean string key. Delegates to the consolidated
    ``railway.ingestion.pl_normalization.norm_pl_key`` (behaviour unchanged, Phase A)."""
    return pl_normalization.norm_pl_key(v)


def _clean_str(v):
    return "" if v is None else str(v).strip()


# ======================================================================
# STRATEGIC DOMAIN  (railways.xlsx)
# ======================================================================
def load_strategic_stock() -> pd.DataFrame:
    """Load the primary 'Stock as on 31.03.2026' sheet into a tidy frame.

    Returns one row per valid PL code with the 12 demand-history fields plus
    the six depot-wise stock columns (used later by AnyLogistix).
    """
    wb = openpyxl.load_workbook(cfg.STRATEGIC_WORKBOOK, read_only=True, data_only=True)
    ws = wb[cfg.STRATEGIC_STOCK_SHEET]
    rows = list(ws.iter_rows(values_only=True))

    num_fields = ["FY2020_21", "FY2022_23", "FY2023_24", "FY2024_25", "FY2025_26",
                  "AAC", "EAR_Qty", "Pending_Supply", "Current_Stock", "Unit_Cost"]

    records = []
    for r in rows[cfg.STRATEGIC_DATA_START_ROW - 1:]:
        def cell(key):
            idx = cfg.STRATEGIC_COLS[key]
            return r[idx] if idx < len(r) else None

        pl = _norm_pl(cell("PL_Code"))
        if pl is None:                       # skip blank / spacer rows
            continue
        rec = {"PL_Code": pl, "Description": _clean_str(cell("Description"))}
        for f in num_fields:
            rec[f] = _to_num(cell(f))
        # depot-wise stock (kept for AnyLogistix; not part of demand_history)
        for depot, idx in cfg.STRATEGIC_DEPOT_COLS.items():
            rec[f"stock_{depot}"] = _to_num(r[idx] if idx < len(r) else None)
        records.append(rec)

    wb.close()
    return pd.DataFrame.from_records(records)


def load_safety_vital() -> pd.DataFrame:
    """Load the Safety/Vital flag (S/V/N) keyed by PL code from the EARAAC sheet."""
    wb = openpyxl.load_workbook(cfg.STRATEGIC_WORKBOOK, read_only=True, data_only=True)
    ws = wb[cfg.EARAAC_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    recs = []
    for r in rows[cfg.EARAAC_DATA_START_ROW - 1:]:
        pl = _norm_pl(r[cfg.EARAAC_COLS["PL_Code"]] if len(r) > cfg.EARAAC_COLS["PL_Code"] else None)
        if pl is None:
            continue
        flag = r[cfg.EARAAC_COLS["Safety_Vital"]] if len(r) > cfg.EARAAC_COLS["Safety_Vital"] else None
        recs.append({"PL_Code": pl, "Safety_Vital": (_clean_str(flag).upper() or None)})
    wb.close()
    return pd.DataFrame.from_records(recs).drop_duplicates(subset="PL_Code")


def load_division_consumption() -> pd.DataFrame:
    """Load division-wise consumption (MAS/TPJ/SA/MDU/PGT/TVC) from 'EAR Consumptions'."""
    wb = openpyxl.load_workbook(cfg.STRATEGIC_WORKBOOK, read_only=True, data_only=True)
    ws = wb[cfg.EAR_CONSUMPTION_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    recs = []
    for r in rows[cfg.EAR_CONSUMPTION_DATA_START_ROW - 1:]:
        pl = _norm_pl(r[cfg.EAR_CONSUMPTION_COLS["PL_Code"]] if len(r) > cfg.EAR_CONSUMPTION_COLS["PL_Code"] else None)
        if pl is None:
            continue
        rec = {"PL_Code": pl,
               "Description": _clean_str(r[cfg.EAR_CONSUMPTION_COLS["Description"]]
                                         if len(r) > cfg.EAR_CONSUMPTION_COLS["Description"] else "")}
        for div, idx in cfg.EAR_DIVISION_CONSUMPTION_COLS.items():
            rec[div] = _to_num(r[idx] if idx < len(r) else None)
        recs.append(rec)
    wb.close()
    return pd.DataFrame.from_records(recs)


def build_demand_history(write: bool = True) -> pd.DataFrame:
    """Assemble railway_demand_history.csv from the strategic stock sheet."""
    df = load_strategic_stock()
    out = df[DEMAND_HISTORY_FIELDS].copy()
    if write:
        cfg.ensure_output_dirs()
        out.to_csv(cfg.DEMAND_HISTORY_CSV, index=False)
    return out


# ======================================================================
# OPERATIONAL DOMAIN  (railway_stock_summary.xlsx -- corrupt stylesheet)
# ======================================================================
def _xlsx_rows_via_xml(path, sheet_xml="xl/worksheets/sheet1.xml"):
    """Yield row dicts {col_index: value} from a single-sheet xlsx via raw OOXML.

    Delegates to the consolidated ``railway.ingestion.excel_reader.iter_sheet_rows``
    (name/signature kept for internal callers; behaviour byte-identical, Phase A).
    """
    yield from excel_reader.iter_sheet_rows(path, sheet_xml)


def load_operational_stock() -> pd.DataFrame:
    """Load railway_stock_summary.xlsx (907 depot items) for the operational pipeline.

    Composite columns are split: PL-Code from 'PL-Code/Type/Usage', and the
    'Last Receipt/Issue Dt.' field into receipt / issue tokens (DD-MM-YYYY).
    """
    rows = list(_xlsx_rows_via_xml(cfg.OPERATIONAL_WORKBOOK))
    recs = []
    for cells in rows[1:]:                          # skip header row
        pl_raw = cells.get(4)
        if not pl_raw:
            continue
        pl = pl_raw.split("/")[0].strip()
        date_raw = cells.get(7) or ""
        parts = date_raw.split("/")
        receipt = parts[0].strip() if len(parts) > 0 else ""
        issue = parts[1].strip() if len(parts) > 1 else ""
        recs.append({
            "PL_Code": pl,
            "Description": _clean_str(cells.get(6)),
            "Current_Stock": _to_num(cells.get(8)),
            "Unit": _clean_str(cells.get(9)),
            "Unit_Cost": _to_num(cells.get(12)),
            "Inventory_Value": _to_num(cells.get(13)),
            "Last_Receipt_Date": receipt or None,
            "Last_Issue_Date": issue or None,
            "Depot": _clean_str(cells.get(2)),          # Consignee Depot (operational lineage)
        })
    return pd.DataFrame.from_records(recs)


# ======================================================================
# VALIDATION  (Step-2 gate)
# ======================================================================
def validate_demand_history(df: pd.DataFrame, safety: pd.DataFrame | None = None) -> dict:
    """Return a structured validation report for the strategic demand history."""
    n = len(df)
    dup_mask = df["PL_Code"].duplicated(keep=False)
    duplicates = sorted(df.loc[dup_mask, "PL_Code"].unique().tolist())

    def n_missing(col):
        return int(df[col].isna().sum()) if df[col].dtype.kind == "f" else int((df[col] == "").sum())

    report = {
        "row_count": n,
        "unique_pl_codes": int(df["PL_Code"].nunique()),
        "pl_is_unique": bool(df["PL_Code"].is_unique),
        "duplicate_pl_codes": duplicates,
        "null_counts": {c: n_missing(c) for c in DEMAND_HISTORY_FIELDS},
        "missing_AAC": int(df["AAC"].isna().sum()),
        "missing_EAR": int(df["EAR_Qty"].isna().sum()),
        "missing_Unit_Cost": int(df["Unit_Cost"].isna().sum()),
        "composite_pl_codes": sorted(df.loc[df["PL_Code"].str.contains("/", na=False), "PL_Code"].tolist()),
    }
    if safety is not None:
        matched = df["PL_Code"].isin(set(safety["PL_Code"])).sum()
        report["safety_flag_matched"] = int(matched)
        report["safety_flag_unmatched"] = int(n - matched)
    # success gate: rows present AND PL codes unique
    report["validation_passed"] = bool(n > 0 and report["pl_is_unique"])
    return report


def _print_report(report: dict, operational_rows: int):
    print("=" * 72)
    print("STEP 2 VALIDATION REPORT  -- railway_demand_history.csv")
    print("=" * 72)
    print(f"Strategic rows (railways.xlsx)        : {report['row_count']}")
    print(f"Unique PL codes                       : {report['unique_pl_codes']}")
    print(f"PL codes unique?                      : {report['pl_is_unique']}")
    print(f"Duplicate PL codes                    : {report['duplicate_pl_codes'] or 'none'}")
    print(f"Composite PL codes (a/b)              : {report['composite_pl_codes'] or 'none'}")
    print(f"Missing AAC                           : {report['missing_AAC']}")
    print(f"Missing EAR_Qty                       : {report['missing_EAR']}")
    print(f"Missing Unit_Cost                     : {report['missing_Unit_Cost']}")
    if "safety_flag_matched" in report:
        print(f"Safety/Vital flag matched / unmatched : "
              f"{report['safety_flag_matched']} / {report['safety_flag_unmatched']}")
    print(f"Operational rows (stock_summary.xlsx) : {operational_rows}")
    print("-" * 72)
    print("Per-field null counts:")
    for c, v in report["null_counts"].items():
        print(f"   {c:16s}: {v}")
    print("-" * 72)
    print(f"VALIDATION PASSED: {report['validation_passed']}")
    print("=" * 72)


def run():
    """Build demand history, load operational stock, print + return validation."""
    df = build_demand_history(write=True)
    safety = load_safety_vital()
    op = load_operational_stock()
    report = validate_demand_history(df, safety)
    _print_report(report, operational_rows=len(op))
    print(f"\nWrote: {cfg.DEMAND_HISTORY_CSV}")
    return df, report


if __name__ == "__main__":
    run()
