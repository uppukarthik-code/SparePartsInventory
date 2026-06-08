"""
One-off ingestion: rebuild raw_data/railway_stock_summary.xlsx (the consolidated
operational workbook the pipeline reads) by concatenating the six per-division
stock_history.xlsx extracts, placing columns at the indices load_operational_stock()
expects (Depot=2, PL=4, Desc=6, Date=7, Stock=8, Unit=9, Unit_Cost=12, Value=13).

Source per-division layout (0-based): #=1, Depot=2, Ledger=3, PL=4, Cat=5, Desc=6,
Date=7, Stock=8, Unit=9, Threshold=10, AvgRate=11, Value=12, Action=13.

Read via raw OOXML (the workbooks carry a stylesheet openpyxl can't parse); written
via a fresh, clean openpyxl workbook. NOT part of the package -- a restore utility.
"""
import os, zipfile
from xml.etree import ElementTree as ET
from openpyxl import Workbook

PROJ = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(PROJ, "raw_data")
DIVS = ["MAS", "SA", "TPJ", "MDU", "PGT", "TVC"]
NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

# source 0-based idx -> target 0-based idx (loader's expected positions)
COLMAP = {2: 2, 4: 4, 6: 6, 7: 7, 8: 8, 9: 9, 11: 12, 12: 13}
NUMERIC_TARGETS = {8, 12, 13}  # Stock, Unit_Cost, Value


def read_rows(path):
    z = zipfile.ZipFile(path)
    shared = []
    if "xl/sharedStrings.xml" in z.namelist():
        sst = ET.fromstring(z.read("xl/sharedStrings.xml"))
        for si in sst.iter(NS + "si"):
            shared.append("".join(t.text or "" for t in si.iter(NS + "t")))

    def ci(ref):
        letters = "".join(c for c in ref if c.isalpha())
        idx = 0
        for ch in letters:
            idx = idx * 26 + (ord(ch) - 64)
        return idx - 1

    sheet = ET.fromstring(z.read("xl/worksheets/sheet1.xml"))
    for row in sheet.iter(NS + "row"):
        cells = {}
        for c in row.findall(NS + "c"):
            t = c.get("t"); v = c.find(NS + "v"); isn = c.find(NS + "is")
            if t == "s" and v is not None:
                val = shared[int(v.text)]
            elif t == "inlineStr" and isn is not None:
                val = "".join(x.text or "" for x in isn.iter(NS + "t"))
            elif v is not None:
                val = v.text
            else:
                val = None
            cells[ci(c.get("r"))] = val
        yield cells


def to_num(x):
    try:
        return float(str(x).replace(",", "").strip())
    except Exception:
        return None


wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# Header row (row 1) -- loader skips it; labels at target positions for readability.
HEADERS = {2: "Consignee Depot", 4: "PL-Code/Type/Usage", 6: "Brief Description",
           7: "Last Receipt / Issue Dt.", 8: "Stock", 9: "Unit",
           12: "Average Rate (Rs.)", 13: "Value (Rs.)", 0: "#", 1: "Business_Unit"}
for tgt, label in HEADERS.items():
    ws.cell(row=1, column=tgt + 1, value=label)

out_row = 2
per_div = {}
for d in DIVS:
    p = os.path.join(RAW, "Railway_Operations", d, "stock_history.xlsx")
    src = list(read_rows(p))
    n = 0
    for cells in src[1:]:                       # skip header
        if not cells.get(4):                    # no PL-Code/Type/Usage -> skip (loader does)
            continue
        ws.cell(row=out_row, column=1, value=None)
        ws.cell(row=out_row, column=2, value=d)   # Business_Unit tag (col B, unused by loader)
        for s_idx, t_idx in COLMAP.items():
            val = cells.get(s_idx)
            if t_idx in NUMERIC_TARGETS:
                num = to_num(val)
                ws.cell(row=out_row, column=t_idx + 1, value=num if num is not None else val)
            else:
                ws.cell(row=out_row, column=t_idx + 1, value=val)
        out_row += 1
        n += 1
    per_div[d] = n

dest = os.path.join(RAW, "railway_stock_summary.xlsx")
wb.save(dest)
total = out_row - 2
print("WROTE:", dest)
print("Rows per division:", per_div)
print("Total operational rows:", total)
